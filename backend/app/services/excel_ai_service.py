import json
import math
import os
import re
import ast
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import get_column_letter

from app.core.config import settings

try:
    from google import genai as google_genai
except ImportError:  # pragma: no cover
    google_genai = None

try:
    from xlcalculator import Evaluator, ModelCompiler
except ImportError:  # pragma: no cover
    ModelCompiler = None
    Evaluator = None


EXCEL_AGENT_SYSTEM_PROMPT = """You are Axel AI — an expert spreadsheet assistant. You receive a DATA SUMMARY and Spreadsheet JSON.

Your job: understand the user's intent, then either (A) answer only in text, or (B) apply changes to the workbook via the JSON plan so results appear in Excel and in the app preview.

Return ONLY strict JSON in this format (no markdown, no text outside the object):
{
  "explanation": "bullet points with leading dash + space on each line",
  "changes": [],
  "charts": [],
  "dashboard": {"create": false},
  "readOnly": true
}

RULE 1 — STRICT JSON ONLY.

RULE 2 — WHEN TO MODIFY THE WORKBOOK (readOnly=false):
Set readOnly=false and populate "changes" (and optionally "charts" / "dashboard") when the user wants ANY of:
- Adding columns, rows, formulas, summaries, pivot-style tables, flags, calculated fields
- Performing operations: sort/filter results materialized on a sheet, group counts, percentages, duplicates markers, new analysis sheets
- Creating charts or dashboards they asked for
- "Do this in the spreadsheet / Excel / file", "apply to my data", "update the sheet", "run these steps"

If the user gives a multi-step numbered list that includes creating columns, summary tables, or formulas — that is a WRITE request: execute it in the workbook AND summarize in "explanation".

RULE 3 — WHEN TO STAY READ-ONLY (readOnly=true):
Set readOnly=true, empty changes/charts, dashboard.create=false ONLY when the user clearly wants information-only with NO file change, for example:
- "How many rows?", "What does column X mean?", "Explain this dataset" without asking to add or change anything
- Quick stats or insights with no request to alter the file

When unsure whether they want the file changed: if they mention adding, creating, calculating, marking, summarizing in the sheet, or "perform operations" — prefer readOnly=false and implement it.

RULE 4 — SAFETY / CELL RULES:
- Do NOT delete or clear existing data unless the user explicitly says delete/remove/clear.
- Prefer NOT to overwrite populated data cells. Add new columns starting at the first FREE column from DATA SUMMARY; add new blocks or summary tables on new sheet(s) when the change would crowd or replace core data.
- Use Excel formulas (=IF, =COUNTIF, =SUM, =AVERAGE, etc.) when the user asks for formulas or derived fields.
- Each change object: {"sheet":"Sheet1","cell":"A1","formula":"=..."} OR {"sheet":"Sheet1","cell":"A1","value":...}. Use "formula" key for formulas (include leading =).
- For visual highlighting, you may add "fillColor":"FFF59D" (hex RGB, with or without leading #) on a change object to color that cell.

RULE 5 — CHARTS:
- If they ask for a chart: add a small summary range with real values or formulas first, then add "charts" with a valid dataRange pointing at that range.
- If user asks for chart/diagram/visualization, charts[] must be non-empty.

RULE 6 — EXPLANATION:
- Always use bullet lines starting with "- " (one bullet per line, newline-separated).
- No markdown bold/asterisks. Never one giant paragraph — one idea per bullet line.
- Describe what you changed in the workbook and key insights in plain language (not raw formula dumps).
- Only mention operations that are actually present in changes/charts/dashboard.
- If user gave numbered tasks, output bullets grouped as Task 1, Task 2, ... (max 1-2 short lines each).

RULE 7 — SCOPE:
Use only sheets/columns/rows present in the provided data unless creating new sheet names you include in "changes".
"""


class ExcelAIService:
    def __init__(self):
        self.client = None
        if settings.GEMINI_API_KEY and google_genai is not None:
            self.client = google_genai.Client(api_key=settings.GEMINI_API_KEY)

    @staticmethod
    def _format_explanation_as_bullets(text: Any) -> str:
        raw = str(text or "").strip()
        if not raw:
            return "- Processed request successfully."
        parsed_struct = None
        if (raw.startswith("[") and raw.endswith("]")) or (raw.startswith("{") and raw.endswith("}")):
            try:
                parsed_struct = json.loads(raw)
            except Exception:
                try:
                    parsed_struct = ast.literal_eval(raw)
                except Exception:
                    parsed_struct = None
        if isinstance(parsed_struct, list):
            items = [str(x).strip() for x in parsed_struct if str(x).strip()]
            if items:
                return "\n".join(f"- {re.sub(r'^[-•*]\\s+', '', it)}" for it in items)
        if isinstance(parsed_struct, dict):
            items = [str(v).strip() for v in parsed_struct.values() if str(v).strip()]
            if items:
                return "\n".join(f"- {re.sub(r'^[-•*]\\s+', '', it)}" for it in items)
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        if not lines:
            return "- Processed request successfully."
        if any(re.match(r"^[-•*]\s+", ln) for ln in lines):
            return "\n".join(re.sub(r"^[-•*]\s+", "- ", ln) for ln in lines)
        if len(lines) > 1:
            return "\n".join(f"- {re.sub(r'^\d+\.\s*', '', ln)}" for ln in lines)
        single = lines[0]
        parts = [p.strip() for p in re.split(r"\s*(?=\d+\.\s+)", single) if p.strip()]
        if len(parts) > 1:
            return "\n".join(f"- {re.sub(r'^\d+\.\s*', '', p)}" for p in parts)
        return f"- {single}"

    @staticmethod
    def _split_sheet_and_range(data_range: str, fallback_sheet: str) -> Tuple[str, str]:
        """
        Accept A1:B10 or Sheet!A1:B10 (including quoted sheet names).
        Returns (sheet_name, plain_a1_range_without_sheet_prefix).
        """
        raw = str(data_range or "").strip()
        if not raw:
            return fallback_sheet, ""
        m = re.match(r"^\s*(?:'([^']+)'|([^!]+))!(\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)\s*$", raw, flags=re.IGNORECASE)
        if m:
            sheet_name = (m.group(1) or m.group(2) or fallback_sheet).strip()
            return sheet_name, m.group(3).replace("$", "")
        return fallback_sheet, raw.replace("$", "")

    @staticmethod
    def _native_xl_value(v: Any) -> Any:
        """Convert xlcalculator / Excel types to JSON-friendly Python values."""
        if v is None:
            return ""
        if isinstance(v, bool):
            return v
        if isinstance(v, (int,)):
            return v
        if isinstance(v, float):
            if math.isnan(v) or math.isinf(v):
                return ""
            return v
        if isinstance(v, str):
            return v
        try:
            from xlcalculator.xlfunctions.func_xltypes import BLANK
            if v is BLANK:
                return ""
        except ImportError:
            pass
        try:
            from xlcalculator import xlerrors
            if isinstance(v, xlerrors.ExcelError):
                return str(v)
        except ImportError:
            pass
        inner = getattr(v, "value", v)
        if isinstance(inner, (str, int, float, bool)):
            if isinstance(inner, float) and (math.isnan(inner) or math.isinf(inner)):
                return ""
            return inner
        return str(v)

    @classmethod
    def enrich_sheet_json_and_changes_from_file(
        cls,
        file_path: str,
        sheet_json: Dict[str, Dict[str, Dict[str, Any]]],
        changes_log: Optional[List[Dict[str, Any]]] = None,
    ) -> Tuple[Dict[str, Dict[str, Dict[str, Any]]], Optional[List[Dict[str, Any]]]]:
        """Evaluate formulas in a saved workbook and merge computed values into preview JSON and change log."""
        if not file_path or not os.path.isfile(file_path) or ModelCompiler is None or Evaluator is None:
            return sheet_json, changes_log
        try:
            mc = ModelCompiler()
            model = mc.read_and_parse_archive(file_path)
            ev = Evaluator(model)
            for addr in list(model.formulae.keys()):
                try:
                    ev.evaluate(addr)
                except Exception:
                    continue
            for sheet_name, cells in sheet_json.items():
                for coord, entry in cells.items():
                    full = f"{sheet_name}!{coord}"
                    if full not in model.formulae:
                        # If the preview contains formula text, never show it.
                        if entry.get("formula"):
                            entry["value"] = ""
                        continue
                    entry["value"] = cls._native_xl_value(model.get_cell_value(full))
            if changes_log:
                for ch in changes_log:
                    after = ch.get("after")
                    full = f"{ch['sheet']}!{ch['cell']}"
                    if isinstance(after, str) and after.startswith("="):
                        ch["after_display"] = cls._native_xl_value(model.get_cell_value(full))
                    else:
                        ch["after_display"] = after
            return sheet_json, changes_log
        except Exception:
            return sheet_json, changes_log

    @staticmethod
    def workbook_to_sheet_json(
        workbook: Workbook,
        max_cells: int = 2000,
        max_rows_per_sheet: int = 200,
        max_cols_per_sheet: int = 40,
    ) -> Dict[str, Dict[str, Dict[str, Any]]]:
        out: Dict[str, Dict[str, Dict[str, Any]]] = {}
        written = 0
        for ws in workbook.worksheets:
            out[ws.title] = {}
            max_row = min(ws.max_row or 1, max_rows_per_sheet)
            max_col = min(ws.max_column or 1, max_cols_per_sheet)
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue
                    is_formula = isinstance(val, str) and val.startswith("=")
                    entry = {
                        # Never display raw formula text in the UI preview.
                        # We keep the formula in `formula` and fill `value` later
                        # after evaluating the workbook (when possible).
                        "value": "" if is_formula else (val if not isinstance(val, bytes) else str(val)),
                        "formula": val if is_formula else None,
                    }
                    out[ws.title][cell.coordinate] = entry
                    written += 1
                    if written >= max_cells:
                        return out
        return out

    @staticmethod
    def _build_data_summary(sheet_json: Dict[str, Any]) -> str:
        """Build a concise data summary for Gemini: columns, row count, sample rows, and free column."""
        parts: List[str] = []
        for sheet_name, cells in sheet_json.items():
            if not cells:
                continue
            max_row = 0
            max_col = 0
            headers: Dict[int, str] = {}
            sample_rows: Dict[int, Dict[int, Any]] = {}
            for cell_ref, cell_data in cells.items():
                m = re.match(r"^([A-Z]+)(\d+)$", cell_ref)
                if not m:
                    continue
                col_idx = 0
                for ch in m.group(1):
                    col_idx = col_idx * 26 + (ord(ch) - 64)
                row_idx = int(m.group(2))
                max_row = max(max_row, row_idx)
                max_col = max(max_col, col_idx)
                val = cell_data.get("value")
                if row_idx == 1:
                    headers[col_idx] = str(val) if val is not None else ""
                elif row_idx <= 4:
                    sample_rows.setdefault(row_idx, {})[col_idx] = val

            col_names = [headers.get(c, f"Col{c}") for c in sorted(headers.keys())]
            free_col = get_column_letter(max_col + 1) if max_col else "H"
            parts.append(
                f"Sheet '{sheet_name}': {max_row - 1} data rows, {max_col} columns.\n"
                f"Columns: {col_names}\n"
                f"Data range: A1:{get_column_letter(max_col)}{max_row}\n"
                f"Free column for helper data: {free_col} onwards"
            )
            for r in sorted(sample_rows.keys()):
                row_vals = [str(sample_rows[r].get(c, "")) for c in sorted(headers.keys())]
                parts.append(f"  Row {r}: {row_vals}")
        return "\n".join(parts) if parts else "Empty spreadsheet."

    @staticmethod
    def _extract_json(text: str) -> Optional[Dict[str, Any]]:
        if not text:
            return None
        candidate = text.strip()
        if candidate.startswith("```"):
            candidate = re.sub(r"^```(?:json)?\s*", "", candidate)
            candidate = re.sub(r"\s*```$", "", candidate).strip()
        try:
            if candidate.startswith("{"):
                parsed = json.loads(candidate)
                if isinstance(parsed, dict):
                    return parsed
        except json.JSONDecodeError:
            pass
        match = re.search(r"(\{[\s\S]*\})", text)
        if not match:
            return None
        try:
            parsed = json.loads(match.group(1))
            return parsed if isinstance(parsed, dict) else None
        except json.JSONDecodeError:
            return None

    def _gemini_generate_json(
        self,
        user_message: str,
        sheet_json: Dict[str, Any],
    ) -> Dict[str, Any]:
        if google_genai is None:
            raise RuntimeError("Gemini SDK is missing. Install the google-genai package in backend environment.")
        if not self.client:
            raise RuntimeError("Gemini is not configured. Please set GEMINI_API_KEY.")
        data_summary = self._build_data_summary(sheet_json)
        prompt = (
            f"{EXCEL_AGENT_SYSTEM_PROMPT}\n\n"
            f"DATA SUMMARY:\n{data_summary}\n\n"
            f"User request:\n{user_message}\n\n"
            "If the user wants the workbook modified (new columns, formulas, extra sheets, summary tables, charts), "
            "you MUST set readOnly=false and include every cell write in the \"changes\" array (and charts/dashboard if requested). "
            "If you only answer facts about the data with no file change, use readOnly=true and empty changes.\n\n"
            f"Spreadsheet JSON:\n{json.dumps(sheet_json, default=str)[:250000]}"
        )
        response = self.client.models.generate_content(
            model=settings.GEMINI_MODEL,
            contents=prompt,
        )
        parsed = self._extract_json(getattr(response, "text", ""))
        if parsed:
            return parsed

        retry_prompt = (
            f"{prompt}\n\nYou did not return valid JSON. RETURN ONLY the JSON object, no markdown, no text."
        )
        retry_response = self.client.models.generate_content(
            model=settings.GEMINI_MODEL,
            contents=retry_prompt,
        )
        parsed_retry = self._extract_json(getattr(retry_response, "text", ""))
        if parsed_retry:
            return parsed_retry
        raise RuntimeError("Gemini returned invalid JSON twice.")

    @staticmethod
    def _normalize_response(raw: Dict[str, Any]) -> Dict[str, Any]:
        changes = raw.get("changes") if isinstance(raw.get("changes"), list) else []
        charts = raw.get("charts") if isinstance(raw.get("charts"), list) else []
        dashboard = raw.get("dashboard") if isinstance(raw.get("dashboard"), dict) else {"create": False}
        read_only = bool(raw.get("readOnly", False))
        if changes or charts or bool(dashboard.get("create")):
            read_only = False
        return {
            "explanation": ExcelAIService._format_explanation_as_bullets(
                raw.get("explanation", "Processed request successfully.")
            ),
            "changes": changes,
            "charts": charts,
            "dashboard": dashboard,
            "readOnly": read_only,
        }

    @staticmethod
    def _ensure_sheet(workbook: Workbook, sheet_name: str):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        return workbook.create_sheet(title=sheet_name)

    @staticmethod
    def _build_chart(chart_cfg: Dict[str, Any], ws) -> Optional[Any]:
        chart_type = str(chart_cfg.get("type", "bar")).lower()
        _, data_range = ExcelAIService._split_sheet_and_range(
            str(chart_cfg.get("dataRange", "")).strip(),
            ws.title,
        )
        if not data_range:
            return None
        try:
            min_col, min_row, max_col, max_row = range_boundaries(data_range)
        except Exception:
            return None
        if max_row <= min_row:
            return None

        if chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()
        chart.title = str(chart_cfg.get("title", "Chart"))

        data_ref = Reference(ws, min_col=min_col + 1 if (max_col - min_col) >= 1 else min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        if (max_col - min_col) >= 1:
            categories = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
            chart.set_categories(categories)
        return chart

    def apply_plan_to_workbook(
        self,
        workbook: Workbook,
        ai_plan: Dict[str, Any],
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any], List[str]]:
        changes_log: List[Dict[str, Any]] = []
        charts_created: List[Dict[str, Any]] = []
        dashboard_summary: Dict[str, Any] = {"created": False, "sheet": None, "elements": 0}
        dashboard_kpis: List[Dict[str, Any]] = []
        touched_sheets: set[str] = set()

        for change in ai_plan.get("changes", []):
            sheet_name = str(change.get("sheet", "Sheet1"))
            cell_ref = str(change.get("cell", "")).strip()
            if not cell_ref:
                continue
            ws = self._ensure_sheet(workbook, sheet_name)
            touched_sheets.add(sheet_name)
            before = ws[cell_ref].value
            if change.get("formula"):
                ws[cell_ref].value = str(change["formula"])
            elif "value" in change:
                ws[cell_ref].value = change.get("value")
            fill_color = str(change.get("fillColor", "")).strip().lstrip("#")
            if fill_color:
                if len(fill_color) == 6:
                    fill_color = f"FF{fill_color.upper()}"
                elif len(fill_color) == 8:
                    fill_color = fill_color.upper()
                if len(fill_color) == 8:
                    ws[cell_ref].fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
            after = ws[cell_ref].value
            changes_log.append({
                "sheet": sheet_name,
                "cell": cell_ref,
                "before": before,
                "after": after,
                "fillColor": str(change.get("fillColor", "")).strip() or None,
            })

        for idx, chart_cfg in enumerate(ai_plan.get("charts", [])):
            sheet_name = str(chart_cfg.get("sheet", "Sheet1"))
            range_sheet, normalized_range = self._split_sheet_and_range(
                str(chart_cfg.get("dataRange", "")).strip(),
                sheet_name,
            )
            source_sheet_name = range_sheet or sheet_name
            source_ws = self._ensure_sheet(workbook, source_sheet_name)
            # User requirement: place all diagrams in the first sheet.
            target_sheet_name = workbook.sheetnames[0] if workbook.sheetnames else source_sheet_name
            target_ws = self._ensure_sheet(workbook, target_sheet_name)
            touched_sheets.add(source_sheet_name)
            touched_sheets.add(target_sheet_name)
            normalized_cfg = {**chart_cfg, "sheet": source_sheet_name, "dataRange": normalized_range}
            chart_obj = self._build_chart(normalized_cfg, source_ws)
            if chart_obj is None:
                continue
            position = str(chart_cfg.get("position", "E5"))
            target_ws.add_chart(chart_obj, position)
            charts_created.append({
                "index": idx,
                "sheet": target_sheet_name,
                "title": str(chart_cfg.get("title", "Chart")),
                "type": str(chart_cfg.get("type", "bar")),
                "dataRange": normalized_range or None,
                "position": position,
            })

        dashboard = ai_plan.get("dashboard", {}) or {}
        if dashboard.get("create"):
            dash_name = str(dashboard.get("sheetName", "Dashboard")).strip() or "Dashboard"
            dash_ws = self._ensure_sheet(workbook, dash_name)
            touched_sheets.add(dash_name)
            dash_ws["A1"] = "AI Dashboard"
            dash_ws["A2"] = "Generated summary metrics and chart references"
            row_idx = 4
            source_sheet = workbook.sheetnames[0]
            for element in dashboard.get("elements", []):
                if element.get("type") == "metric":
                    label = str(element.get("label", "Metric"))
                    value_cell = str(element.get("valueCell", "")).strip()
                    value_formula = str(element.get("valueFormula", "")).strip()
                    dash_ws[f"A{row_idx}"] = label

                    if value_formula:
                        source_ws = self._ensure_sheet(workbook, source_sheet)
                        source_ws[value_cell] = value_formula
                        if value_cell:
                            if "!" in value_cell:
                                dash_ws[f"B{row_idx}"] = f"={value_cell}"
                            else:
                                dash_ws[f"B{row_idx}"] = f"={source_sheet}!{value_cell}"
                    elif value_cell:
                        if "!" in value_cell:
                            dash_ws[f"B{row_idx}"] = f"={value_cell}"
                        else:
                            dash_ws[f"B{row_idx}"] = f"={source_sheet}!{value_cell}"

                    dashboard_kpis.append({
                        "label": label,
                        "formula": value_formula or f"={source_sheet}!{value_cell}" if value_cell else "",
                        "cell": f"B{row_idx}",
                    })
                    row_idx += 1
                elif element.get("type") == "chart":
                    chart_ref_idx = int(element.get("chartRef", -1))
                    if 0 <= chart_ref_idx < len(ai_plan.get("charts", [])):
                        chart_cfg = ai_plan["charts"][chart_ref_idx]
                        source_ws = self._ensure_sheet(workbook, str(chart_cfg.get("sheet", source_sheet)))
                        dash_chart = self._build_chart(chart_cfg, source_ws)
                        if dash_chart is not None:
                            dash_ws.add_chart(dash_chart, f"D{max(4, row_idx)}")
                            row_idx += 12
            dashboard_summary = {
                "created": True,
                "sheet": dash_name,
                "elements": len(dashboard.get("elements", [])),
                "kpis": dashboard_kpis,
            }

        return changes_log, charts_created, dashboard_summary, sorted(touched_sheets)

    @staticmethod
    def _normalize_text(v: Any) -> str:
        return str(v or "").strip().lower()

    @staticmethod
    def _find_header_col(ws, aliases: List[str]) -> Optional[int]:
        alias_norm = [a.strip().lower() for a in aliases]
        for c in range(1, (ws.max_column or 1) + 1):
            h = ExcelAIService._normalize_text(ws.cell(row=1, column=c).value)
            if not h:
                continue
            if any(a == h or a in h for a in alias_norm):
                return c
        return None

    @staticmethod
    def _is_visual_dashboard_query(user_message: str) -> bool:
        q = (user_message or "").lower()
        tokens = ("dashboard", "chart", "pie", "bar", "visual", "visualization", "diagram", "top 5", "top five")
        return any(t in q for t in tokens)

    def _apply_visual_dashboard_fallback(
        self,
        workbook: Workbook,
        user_message: str,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any], List[str], str]:
        """
        Deterministic visual fallback for locality + nearby-shops dashboard requests.
        Ensures charts exist in workbook/download and can be previewed from sheet_json.
        """
        if not workbook.sheetnames:
            return [], [], {"created": False, "sheet": None, "elements": 0}, [], "- No data sheet found to create visuals."
        src_ws = workbook[workbook.sheetnames[0]]
        area_col = self._find_header_col(src_ws, ["area/locality", "area", "locality", "city", "location"])
        shops_col = self._find_header_col(src_ws, ["nearby shop", "nearby shop 1", "nearby shop1", "shop", "xerox shop", "nearby xerox"])
        college_col = self._find_header_col(src_ws, ["college", "college name"])
        if area_col is None:
            return [], [], {"created": False, "sheet": None, "elements": 0}, [], "- Could not detect area/locality column for visual dashboard."

        area_counts: Dict[str, int] = {}
        shop_freq: Dict[str, int] = {}
        rows = max(src_ws.max_row or 1, 1)
        for r in range(2, rows + 1):
            area = str(src_ws.cell(row=r, column=area_col).value or "").strip()
            if area:
                area_counts[area] = area_counts.get(area, 0) + 1
            if shops_col is not None:
                shop = str(src_ws.cell(row=r, column=shops_col).value or "").strip()
                if shop:
                    shop_freq[shop] = shop_freq.get(shop, 0) + 1

        if not area_counts:
            return [], [], {"created": False, "sheet": None, "elements": 0}, [], "- No area/locality values found for creating visuals."

        changes_log: List[Dict[str, Any]] = []
        charts_created: List[Dict[str, Any]] = []
        touched_sheets = {src_ws.title}
        start_col = (src_ws.max_column or 1) + 2

        def addr(col_offset: int, row_idx: int) -> str:
            return f"{get_column_letter(start_col + col_offset)}{row_idx}"

        def write(col_offset: int, row_idx: int, value: Any):
            cell_ref = addr(col_offset, row_idx)
            before = src_ws[cell_ref].value
            src_ws[cell_ref].value = value
            changes_log.append({"sheet": src_ws.title, "cell": cell_ref, "before": before, "after": value, "fillColor": None})

        write(0, 1, "Dashboard Analysis")
        write(0, 2, "Auto-generated visual summary for locality and xerox opportunity")
        write(0, 3, "Charts are placed below. Scroll slightly if needed.")

        sorted_areas = sorted(area_counts.items(), key=lambda x: x[1], reverse=True)
        top5_areas = sorted_areas[:5]
        sorted_shops = sorted(shop_freq.items(), key=lambda x: x[1], reverse=True)[:10]

        # Table 1: colleges per area
        write(0, 4, "Area/Locality")
        write(1, 4, "Number of Colleges")
        row_ptr = 5
        for area, count in sorted_areas[:20]:
            write(0, row_ptr, area)
            write(1, row_ptr, count)
            row_ptr += 1
        table1_end = row_ptr - 1

        # Table 2: nearby shop frequency
        write(3, 4, "Nearby Shop")
        write(4, 4, "Frequency")
        row_ptr2 = 5
        for shop, count in sorted_shops:
            write(3, row_ptr2, shop)
            write(4, row_ptr2, count)
            row_ptr2 += 1
        table2_end = row_ptr2 - 1
        if table2_end < 5:
            write(3, 5, "No nearby shop data")
            write(4, 5, 0)
            table2_end = 5

        # Table 3: top 5 areas
        write(6, 4, "Top 5 Areas")
        write(7, 4, "College Count")
        row_ptr3 = 5
        for area, count in top5_areas:
            write(6, row_ptr3, area)
            write(7, row_ptr3, count)
            row_ptr3 += 1
        table3_end = max(row_ptr3 - 1, 5)

        table1_range = f"{addr(0,4)}:{addr(1,table1_end)}"
        table2_range = f"{addr(3,4)}:{addr(4,table2_end)}"
        table3_range = f"{addr(6,4)}:{addr(7,table3_end)}"
        pos1 = addr(0, 8)
        pos2 = addr(8, 8)
        pos3 = addr(0, 24)
        chart_specs = [
            {
                "sheet": src_ws.title,
                "type": "bar",
                "title": "Number of Colleges per Area/Locality",
                "dataRange": table1_range,
                "position": pos1,
            },
            {
                "sheet": src_ws.title,
                "type": "bar",
                "title": "Frequency of Nearby Shops per College",
                "dataRange": table2_range,
                "position": pos2,
            },
            {
                "sheet": src_ws.title,
                "type": "pie",
                "title": "Top 5 Areas by College Count",
                "dataRange": table3_range,
                "position": pos3,
            },
        ]
        for idx, cfg in enumerate(chart_specs):
            ch_obj = self._build_chart(cfg, src_ws)
            if ch_obj is None:
                continue
            src_ws.add_chart(ch_obj, cfg["position"])
            charts_created.append({
                "index": idx,
                "sheet": cfg["sheet"],
                "title": cfg["title"],
                "type": cfg["type"],
                "dataRange": cfg["dataRange"],
                "position": cfg["position"],
            })

        # Improve first-open readability in downloaded Excel.
        src_ws.sheet_view.zoomScale = 90
        src_ws.freeze_panes = addr(0, 5)
        for off, width in {
            0: 34, 1: 18, 2: 4,
            3: 36, 4: 14, 5: 4,
            6: 32, 7: 16, 8: 4,
            9: 18, 10: 18,
        }.items():
            src_ws.column_dimensions[get_column_letter(start_col + off)].width = width

        top_area_text = ", ".join([f"{a} ({c})" for a, c in top5_areas[:3]]) if top5_areas else "N/A"
        explanation = "\n".join([
            "- Added summary tables directly on Sheet1 so all operations stay in one sheet.",
            "- Added bar chart for number of colleges per area/locality on Sheet1.",
            "- Added bar chart for frequency of nearby xerox shops on Sheet1.",
            "- Added pie chart for top 5 areas with highest college count on Sheet1.",
            f"- Highest opportunity areas based on college density: {top_area_text}.",
        ])
        dashboard_summary = {
            "created": True,
            "sheet": src_ws.title,
            "elements": len(charts_created),
            "kpis": [],
        }
        return changes_log, charts_created, dashboard_summary, sorted(touched_sheets), explanation

    @staticmethod
    def extract_sheet_preview(
        workbook: Workbook,
        sheet_name: Optional[str] = None,
        max_rows: int = 15,
    ) -> Dict[str, Any]:
        """Extract a table preview (columns + rows) from a workbook sheet."""
        if sheet_name and sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        elif workbook.sheetnames:
            ws = workbook[workbook.sheetnames[0]]
            sheet_name = workbook.sheetnames[0]
        else:
            return {"sheet": "", "columns": [], "rows": []}

        max_col = min(ws.max_column or 1, 20)
        max_row_actual = min(ws.max_row or 1, max_rows + 1)

        headers: List[str] = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=1, column=c).value
            headers.append(str(val) if val is not None else f"Column {c}")

        rows: List[Dict[str, Any]] = []
        for r in range(2, max_row_actual + 1):
            row_data: Dict[str, Any] = {}
            non_empty = False
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and val.startswith("="):
                    # Never display raw formula text in any previews.
                    val = ""
                row_data[headers[c - 1]] = val if val is not None else ""
                if val is not None and val != "":
                    non_empty = True
            if non_empty:
                rows.append(row_data)

        return {"sheet": sheet_name, "columns": headers, "rows": rows}

    def process_excel_request(
        self,
        workbook: Workbook,
        user_message: str,
        input_sheet_json: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        sheet_json = input_sheet_json or self.workbook_to_sheet_json(workbook)
        raw = self._gemini_generate_json(user_message=user_message, sheet_json=sheet_json)
        plan = self._normalize_response(raw)

        if plan["readOnly"]:
            return {
                **plan,
                "change_tracking": [],
                "applied_charts": [],
                "dashboard_summary": {"created": False, "sheet": None, "elements": 0},
            }

        change_tracking, applied_charts, dashboard_summary, changed_sheets = self.apply_plan_to_workbook(workbook, plan)
        if self._is_visual_dashboard_query(user_message):
            fb_changes, fb_charts, fb_dash, fb_sheets, fb_explanation = self._apply_visual_dashboard_fallback(
                workbook=workbook,
                user_message=user_message,
            )
            if fb_charts:
                change_tracking = fb_changes
                applied_charts = fb_charts
                dashboard_summary = fb_dash
                changed_sheets = fb_sheets
                plan["explanation"] = self._format_explanation_as_bullets(fb_explanation)
        return {
            **plan,
            "change_tracking": change_tracking,
            "applied_charts": applied_charts,
            "dashboard_summary": dashboard_summary,
            "changed_sheets": changed_sheets,
        }
