import pandas as pd
import io
import os
import numpy as np
import ast
from typing import List, Dict, Any, Optional, Tuple
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.feature_extraction.text import TfidfVectorizer
import json
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

class DataService:
    _BANNED_NAMES = {
        "__import__", "eval", "exec", "open", "compile", "input", "globals", "locals",
        "vars", "dir", "getattr", "setattr", "delattr", "help", "breakpoint",
    }
    _BANNED_ATTR_PREFIXES = ("__",)
    _BANNED_MODULE_PREFIXES = (
        "os", "sys", "subprocess", "pathlib", "shutil", "socket", "requests", "http", "urllib"
    )
    _SAFE_BUILTINS = {
        "len": len,
        "min": min,
        "max": max,
        "sum": sum,
        "abs": abs,
        "round": round,
        "sorted": sorted,
        "list": list,
        "dict": dict,
        "set": set,
        "tuple": tuple,
        "str": str,
        "int": int,
        "float": float,
        "bool": bool,
        "enumerate": enumerate,
        "range": range,
        "zip": zip,
        "any": any,
        "all": all,
        "print": print,
    }

    @staticmethod
    def load_dataset(file_path: str, sheet_name: Optional[str] = None) -> Any:
        """Load dataset. For Excel: returns dict {sheet_name: df}. For CSV: returns single DataFrame."""
        if file_path.endswith('.csv'):
            # Prefer Arrow CSV engine when available for faster parsing on large files.
            try:
                import pyarrow  # noqa: F401
                return pd.read_csv(file_path, engine="pyarrow")
            except Exception:
                return pd.read_csv(file_path, low_memory=False)
        elif file_path.endswith(('.xls', '.xlsx')):
            xl = pd.ExcelFile(file_path)
            if sheet_name:
                return xl.parse(sheet_name)
            return {name: xl.parse(name) for name in xl.sheet_names}
        else:
            raise ValueError("Unsupported file format. Use .csv, .xlsx, or .xls")

    @staticmethod
    def _safe_formula(df: pd.DataFrame, formula: str) -> Optional[pd.Series]:
        """Evaluate simple formula like 'Price * Quantity'. No code execution - only pandas ops."""
        formula = formula.strip()
        allowed = set(df.columns)
        if not allowed:
            return None
        # Only allow: column, operator, column
        ops = {"*": lambda a, b: a * b, "+": lambda a, b: a + b, "-": lambda a, b: a - b,
               "/": lambda a, b: a / b.replace(0, np.nan)}
        for op_str, op_fn in ops.items():
            if op_str in formula:
                parts = formula.split(op_str, 1)
                if len(parts) == 2:
                    left, right = parts[0].strip(), parts[1].strip()
                    if left in allowed and right in allowed:
                        return op_fn(df[left], df[right])
                    break
        return None

    @staticmethod
    def get_single_df(data: Any, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """Extract single DataFrame from load_dataset result."""
        if isinstance(data, pd.DataFrame):
            return data
        if isinstance(data, dict):
            name = sheet_name or list(data.keys())[0]
            return data[name]
        raise ValueError("Invalid data format")

    @staticmethod
    def profile_dataset(df: pd.DataFrame) -> Dict[str, Any]:
        """
        Generate comprehensive metadata about the dataset.
        """
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        text_cols = df.select_dtypes(include=['object']).columns.tolist()
        categorical_cols = [col for col in text_cols if df[col].nunique() < 50]
        
        profile = {
            "total_rows": int(df.shape[0]),
            "total_columns": int(df.shape[1]),
            "columns": list(df.columns),
            "dtypes": df.dtypes.astype(str).to_dict(),
            "missing_values": df.isnull().sum().to_dict(),
            "missing_percentage": (df.isnull().sum() / len(df) * 100).to_dict() if len(df) > 0 else {},
            "shape": df.shape,
            "head": df.head(10).to_dict(orient='records'),
            "numeric_columns": numeric_cols,
            "text_columns": text_cols,
            "categorical_columns": categorical_cols,
            "duplicate_rows": int(df.duplicated().sum()),
            "constant_columns": [col for col in df.columns if df[col].nunique() <= 1],
            "summary": {}
        }
        
        # Add statistical summary for numeric columns
        if numeric_cols:
            profile["summary"]["numeric"] = df[numeric_cols].describe().to_dict()
        
        # Add value counts for categorical columns
        if categorical_cols:
            profile["summary"]["categorical"] = {
                col: df[col].value_counts().head(10).to_dict() 
                for col in categorical_cols[:5]  # Limit to first 5 to avoid huge output
            }
        
        return DataService._sanitize_for_json(profile)

    @staticmethod
    def _sanitize_for_json(data: Any) -> Any:
        """Ensure all values are JSON-compliant (no NaN, Inf, pd.NA, etc.)."""
        if isinstance(data, dict):
            result = {}
            for k, v in data.items():
                # Keys can be NaN from value_counts; ensure JSON-safe keys
                try:
                    if isinstance(k, (np.floating, float)) and (np.isnan(k) or np.isinf(k)):
                        safe_key = "__null__"
                    elif hasattr(pd, 'isna') and pd.isna(k):
                        safe_key = "__null__"
                    elif isinstance(k, (str, int, bool)):
                        safe_key = k
                    else:
                        safe_key = str(k)
                except (TypeError, ValueError):
                    safe_key = "__null__"
                result[safe_key] = DataService._sanitize_for_json(v)
            return result
        elif isinstance(data, list):
            return [DataService._sanitize_for_json(v) for v in data]
        elif isinstance(data, (np.integer,)):
            return int(data)
        elif isinstance(data, (np.floating,)):
            if np.isnan(data) or np.isinf(data):
                return None
            return float(data)
        elif isinstance(data, float):
            if np.isnan(data) or np.isinf(data):
                return None
            return data
        elif isinstance(data, (pd.Timestamp,)):
            return str(data) if pd.notna(data) else None
        elif hasattr(data, 'isoformat'):  # datetime, date
            return data.isoformat() if data is not None else None
        elif data is pd.NA or (hasattr(pd, 'NA') and data is pd.NA):
            return None
        elif isinstance(data, (np.ndarray,)):
            return DataService._sanitize_for_json(data.tolist())
        return data

    @staticmethod
    def _validate_generated_code(code: str) -> None:
        """
        Validate generated Python code before execution.
        Allows dataframe transformations only; blocks imports/system calls.
        """
        if not isinstance(code, str) or not code.strip():
            raise ValueError("Generated code is empty.")

        try:
            tree = ast.parse(code, mode="exec")
        except SyntaxError as exc:
            raise ValueError(f"Generated code has invalid syntax: {exc}") from exc

        allowed_nodes = (
            ast.Module, ast.Expr, ast.Assign, ast.AugAssign, ast.AnnAssign,
            ast.Name, ast.Load, ast.Store, ast.Constant, ast.List, ast.Tuple, ast.Dict, ast.Set,
            ast.Subscript, ast.Slice, ast.Index,
            ast.Attribute, ast.Call, ast.keyword,
            ast.BinOp, ast.UnaryOp, ast.BoolOp, ast.Compare,
            ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Mod, ast.Pow, ast.FloorDiv,
            ast.USub, ast.UAdd, ast.Not, ast.And, ast.Or,
            ast.Eq, ast.NotEq, ast.Lt, ast.LtE, ast.Gt, ast.GtE, ast.In, ast.NotIn, ast.Is, ast.IsNot,
            ast.If, ast.For, ast.While, ast.Break, ast.Continue, ast.Pass,
            ast.ListComp, ast.DictComp, ast.SetComp, ast.GeneratorExp, ast.comprehension,
            ast.IfExp,
            ast.Try, ast.ExceptHandler, ast.Raise,
            ast.Lambda,
            ast.JoinedStr, ast.FormattedValue,
            ast.FunctionDef, ast.arguments, ast.arg, ast.Return,
        )

        for node in ast.walk(tree):
            if not isinstance(node, allowed_nodes):
                raise ValueError(f"Unsupported code construct: {type(node).__name__}")

            if isinstance(node, (ast.Import, ast.ImportFrom)):
                raise ValueError("Import statements are not allowed in generated code.")

            if isinstance(node, ast.Name):
                if node.id in DataService._BANNED_NAMES:
                    raise ValueError(f"Use of '{node.id}' is not allowed.")

            if isinstance(node, ast.Attribute):
                if any(str(node.attr).startswith(prefix) for prefix in DataService._BANNED_ATTR_PREFIXES):
                    raise ValueError(f"Unsafe attribute access: {node.attr}")

            if isinstance(node, ast.Call):
                fn_name = None
                if isinstance(node.func, ast.Name):
                    fn_name = node.func.id
                elif isinstance(node.func, ast.Attribute):
                    fn_name = node.func.attr
                    root = node.func.value
                    if isinstance(root, ast.Name):
                        root_name = root.id
                        if root_name in DataService._BANNED_MODULE_PREFIXES:
                            raise ValueError(f"Module '{root_name}' is not allowed.")
                if fn_name in DataService._BANNED_NAMES:
                    raise ValueError(f"Function '{fn_name}' is not allowed.")

    @staticmethod
    def _strip_disallowed_lines(code: str) -> str:
        """
        Remove unsupported markdown/import lines from model output before AST validation.
        """
        cleaned_lines: List[str] = []
        for line in (code or "").splitlines():
            stripped = line.strip()
            if not stripped:
                cleaned_lines.append(line)
                continue
            if stripped.startswith("```"):
                continue
            if stripped.startswith("import ") or stripped.startswith("from "):
                continue
            cleaned_lines.append(line)
        return "\n".join(cleaned_lines).strip()

    @staticmethod
    def execute_generated_code(df: pd.DataFrame, code: str) -> Tuple[pd.DataFrame, List[str], Dict[str, pd.DataFrame]]:
        """
        Execute validated LLM-generated dataframe transformation code.
        Code must transform variable `df` and keep it as a DataFrame.
        """
        code = DataService._strip_disallowed_lines(code)
        DataService._validate_generated_code(code)

        local_env: Dict[str, Any] = {
            "df": df.copy(),
            "pd": pd,
            "np": np,
            # Optional output contract for advanced Excel operations:
            # generated code can assign additional sheets as {"SheetName": dataframe}
            "excel_outputs": {},
        }
        global_env: Dict[str, Any] = {"__builtins__": DataService._SAFE_BUILTINS}

        exec(compile(code, "<generated_transform>", "exec"), global_env, local_env)

        result = local_env.get("df")
        if not isinstance(result, pd.DataFrame):
            raise ValueError("Generated code did not produce a valid DataFrame in variable 'df'.")

        excel_outputs_raw = local_env.get("excel_outputs", {})
        excel_outputs: Dict[str, pd.DataFrame] = {}
        if isinstance(excel_outputs_raw, dict):
            for sheet_name, sheet_df in excel_outputs_raw.items():
                if isinstance(sheet_name, str) and isinstance(sheet_df, pd.DataFrame):
                    safe_name = (sheet_name or "").strip()[:31] or "Sheet"
                    excel_outputs[safe_name] = sheet_df

        return result, ["Applied generated transformation code from natural language query."], excel_outputs

    @staticmethod
    def _add_dashboard_charts(wb: Workbook) -> None:
        """
        Add Excel-native charts for generated dashboard sheets.
        This ensures downloaded .xlsx files include visible diagrams.
        """
        if "KPI_Summary" in wb.sheetnames:
            ws_kpi = wb["KPI_Summary"]
            if (ws_kpi.max_row or 1) >= 3 and (ws_kpi.max_column or 1) >= 2:
                pie = PieChart()
                pie.title = "KPI Distribution"
                labels = Reference(ws_kpi, min_col=1, min_row=2, max_row=ws_kpi.max_row)
                data = Reference(ws_kpi, min_col=2, min_row=1, max_row=ws_kpi.max_row)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                ws_kpi.add_chart(pie, "D2")

        if "Top_Categories" in wb.sheetnames:
            ws_top = wb["Top_Categories"]
            if (ws_top.max_row or 1) >= 3 and (ws_top.max_column or 1) >= 2:
                bar = BarChart()
                bar.type = "col"
                bar.style = 10
                bar.title = "Top Categories Count"
                bar.y_axis.title = "Count"
                bar.x_axis.title = "Category"
                labels = Reference(ws_top, min_col=1, min_row=2, max_row=ws_top.max_row)
                data = Reference(ws_top, min_col=2, min_row=1, max_row=ws_top.max_row)
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(labels)
                ws_top.add_chart(bar, "D2")

    @staticmethod
    def save_excel_with_outputs(
        output_path: str,
        main_df: pd.DataFrame,
        extra_sheets: Optional[Dict[str, pd.DataFrame]] = None,
    ) -> None:
        """
        Save the transformed dataframe plus optional extra generated tables/sheets to one workbook.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "TransformedData"

        def write_df_to_ws(target_ws, data: pd.DataFrame):
            rows = [list(data.columns)] + data.fillna("").astype(object).values.tolist()
            for row in rows:
                target_ws.append(row)
            if data.shape[0] > 0 and data.shape[1] > 0:
                end_col = get_column_letter(data.shape[1])
                end_row = data.shape[0] + 1
                table_ref = f"A1:{end_col}{end_row}"
                table_name = f"Tbl{target_ws.title[:20].replace(' ', '')}"[:31]
                table = Table(displayName=table_name, ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                target_ws.add_table(table)

        write_df_to_ws(ws, main_df)

        for raw_name, sheet_df in (extra_sheets or {}).items():
            if not isinstance(sheet_df, pd.DataFrame):
                continue
            sheet_name = (raw_name or "GeneratedTable").strip()[:31]
            if not sheet_name:
                sheet_name = "GeneratedTable"
            base = sheet_name
            idx = 1
            while sheet_name in wb.sheetnames:
                suffix = str(idx)
                sheet_name = f"{base[:31 - len(suffix)]}{suffix}"
                idx += 1
            extra_ws = wb.create_sheet(title=sheet_name)
            write_df_to_ws(extra_ws, sheet_df)

        # Embed charts when dashboard/helper sheets are present.
        DataService._add_dashboard_charts(wb)
        wb.save(output_path)

    @staticmethod
    def execute_plan(df: pd.DataFrame, plan: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Execute a list of operations on the DataFrame deterministically.
        Enhanced with more transformation operations.
        """
        df_processed = df.copy()
        execution_log = []

        for idx, raw_op in enumerate(plan):
            # Be defensive: accept dict ops, or skip invalid formats
            if isinstance(raw_op, dict):
                op = raw_op
            elif isinstance(raw_op, (list, tuple)) and len(raw_op) >= 1 and isinstance(raw_op[0], str):
                op = {"type": raw_op[0], "params": raw_op[1] if len(raw_op) > 1 and isinstance(raw_op[1], dict) else {}}
            else:
                execution_log.append(f"Skipped invalid operation at step {idx + 1}: {type(raw_op).__name__}")
                continue

            op_type = op.get("type")
            params = op.get("params", {}) or {}

            try:
                if op_type == "drop_columns":
                    cols = params.get("columns", [])
                    df_processed.drop(columns=cols, errors='ignore', inplace=True)
                    execution_log.append(f"Dropped columns: {cols}")
                
                elif op_type == "drop_constant_columns":
                    constant_cols = [col for col in df_processed.columns if df_processed[col].nunique() <= 1]
                    df_processed.drop(columns=constant_cols, errors='ignore', inplace=True)
                    execution_log.append(f"Dropped constant columns: {constant_cols}")
                
                elif op_type == "drop_na":
                    cols = params.get("columns")
                    how = params.get("how", "any")
                    rows_before = len(df_processed)
                    if cols:
                        df_processed.dropna(subset=cols, how=how, inplace=True)
                    else:
                        df_processed.dropna(how=how, inplace=True)
                    rows_removed = rows_before - len(df_processed)
                    execution_log.append(f"Removed {rows_removed} rows with missing values")

                elif op_type == "fill_na":
                    cols = params.get("columns", [])
                    val = params.get("value")
                    method = params.get("method")
                    
                    if not cols:
                        cols = df_processed.columns.tolist()

                    if val is not None:
                        df_processed[cols] = df_processed[cols].fillna(val)
                        execution_log.append(f"Filled missing values in {cols} with {val}")
                    elif method in ['mean', 'median', 'mode']:
                        for col in cols:
                            if col in df_processed.select_dtypes(include=['number']).columns:
                                if method == 'mean':
                                    fill_val = df_processed[col].mean()
                                elif method == 'median':
                                    fill_val = df_processed[col].median()
                                df_processed[col] = df_processed[col].fillna(fill_val)
                            elif method == 'mode':
                                if not df_processed[col].mode().empty:
                                    fill_val = df_processed[col].mode()[0]
                                    df_processed[col] = df_processed[col].fillna(fill_val)
                        execution_log.append(f"Filled missing values using {method} method")

                elif op_type == "rename_columns":
                    mapping = params.get("mapping", {})
                    df_processed.rename(columns=mapping, inplace=True)
                    execution_log.append(f"Renamed columns: {mapping}")

                elif op_type == "convert_type":
                    cols = params.get("columns", [])
                    target = params.get("target_type")
                    for col in cols:
                        if col in df_processed.columns:
                            if target == 'int':
                                df_processed[col] = pd.to_numeric(df_processed[col], errors='coerce').fillna(0).astype(int)
                            elif target == 'float':
                                df_processed[col] = pd.to_numeric(df_processed[col], errors='coerce')
                            elif target == 'datetime':
                                df_processed[col] = pd.to_datetime(df_processed[col], errors='coerce')
                            elif target == 'str':
                                df_processed[col] = df_processed[col].astype(str)
                    execution_log.append(f"Converted {cols} to {target}")
                
                elif op_type == "remove_duplicates":
                    subset = params.get("subset")
                    rows_before = len(df_processed)
                    df_processed.drop_duplicates(subset=subset, inplace=True)
                    rows_removed = rows_before - len(df_processed)
                    execution_log.append(f"Removed {rows_removed} duplicate rows")
                
                elif op_type == "encode_categorical":
                    cols = params.get("columns", [])
                    for col in cols:
                        if col in df_processed.columns:
                            le = LabelEncoder()
                            df_processed[col] = le.fit_transform(df_processed[col].astype(str))
                    execution_log.append(f"Label encoded categorical columns: {cols}")
                
                elif op_type == "extract_datetime_features":
                    cols = params.get("columns", [])
                    for col in cols:
                        if col in df_processed.columns:
                            df_processed[col] = pd.to_datetime(df_processed[col], errors='coerce')
                            df_processed[f"{col}_year"] = df_processed[col].dt.year
                            df_processed[f"{col}_month"] = df_processed[col].dt.month
                            df_processed[f"{col}_day"] = df_processed[col].dt.day
                            df_processed[f"{col}_dayofweek"] = df_processed[col].dt.dayofweek
                    execution_log.append(f"Extracted datetime features from: {cols}")
                
                elif op_type == "vectorize_text":
                    cols = params.get("columns", [])
                    max_features = params.get("max_features", 50)
                    for col in cols:
                        if col in df_processed.columns:
                            vectorizer = TfidfVectorizer(max_features=max_features, stop_words='english')
                            text_data = df_processed[col].fillna('').astype(str)
                            vectors = vectorizer.fit_transform(text_data)
                            feature_names = [f"{col}_tfidf_{i}" for i in range(vectors.shape[1])]
                            df_vectors = pd.DataFrame(vectors.toarray(), columns=feature_names, index=df_processed.index)
                            df_processed = pd.concat([df_processed.drop(columns=[col]), df_vectors], axis=1)
                    execution_log.append(f"Vectorized text columns: {cols}")
                
                elif op_type == "normalize_numeric":
                    cols = params.get("columns", [])
                    numeric_cols = [col for col in cols if col in df_processed.select_dtypes(include=[np.number]).columns]
                    if numeric_cols:
                        scaler = StandardScaler()
                        df_processed[numeric_cols] = scaler.fit_transform(df_processed[numeric_cols])
                    execution_log.append(f"Normalized numeric columns: {numeric_cols}")

                elif op_type == "slice":
                    rows = params.get("rows", 5)
                    method = params.get("method", "head")
                    rows_before = len(df_processed)
                    
                    if method == "head":
                        df_processed = df_processed.head(rows)
                    elif method == "tail":
                        df_processed = df_processed.tail(rows)
                    elif method == "random":
                        df_processed = df_processed.sample(n=min(rows, len(df_processed)))
                    
                    execution_log.append(f"Sliced dataset to {len(df_processed)} rows using {method}")

                elif op_type == "keep_unique_rows":
                    cols = params.get("columns", [])
                    count = params.get("count", 5)
                    keep = params.get("keep", "first")
                    if not isinstance(cols, list):
                        cols = []
                    valid_cols = [c for c in cols if c in df_processed.columns]
                    try:
                        count = int(count)
                    except (TypeError, ValueError):
                        count = 5
                    count = max(1, count)
                    keep = "last" if keep == "last" else "first"

                    if valid_cols:
                        deduped = df_processed.dropna(subset=valid_cols).drop_duplicates(subset=valid_cols, keep=keep)
                        df_processed = deduped.head(count)
                        execution_log.append(
                            f"Kept {len(df_processed)} unique rows by {valid_cols} (requested {count}, keep={keep})"
                        )
                    else:
                        deduped = df_processed.drop_duplicates(keep=keep)
                        df_processed = deduped.head(count)
                        execution_log.append(
                            f"Kept {len(df_processed)} unique rows across all columns (requested {count}, keep={keep})"
                        )

                elif op_type == "sort":
                    cols = params.get("columns", [])
                    ascending = params.get("ascending", False)
                    if cols:
                        valid_cols = [c for c in cols if c in df_processed.columns]
                        if valid_cols:
                            df_processed = df_processed.sort_values(by=valid_cols, ascending=ascending)
                            execution_log.append(f"Sorted by {valid_cols} ({'ascending' if ascending else 'descending'})")

                elif op_type == "add_column":
                    new_col = params.get("new_column") or params.get("column")
                    formula = params.get("formula") or params.get("expression", "")
                    if new_col and formula:
                        # Safe formula: only column names + * - / (no code execution)
                        result = DataService._safe_formula(df_processed, formula)
                        if result is not None:
                            df_processed[new_col] = result
                            execution_log.append(f"Added column '{new_col}' = {formula}")
                        else:
                            execution_log.append(f"Could not add column '{new_col}': invalid formula")
                    elif new_col and params.get("value") is not None:
                        df_processed[new_col] = params["value"]
                        execution_log.append(f"Added column '{new_col}' with constant value")

                elif op_type == "reorder_columns":
                    order = params.get("column_order", params.get("columns", []))
                    if order:
                        valid = [c for c in order if c in df_processed.columns]
                        extra = [c for c in df_processed.columns if c not in valid]
                        df_processed = df_processed[valid + extra]
                        execution_log.append(f"Reordered columns: {valid[:10]}{'...' if len(valid) > 10 else ''}")

                elif op_type == "move_column":
                    col = params.get("column") or (params.get("columns", [None])[0] if params.get("columns") else None)
                    after = params.get("after_column")
                    if col and col in df_processed.columns and after and after in df_processed.columns:
                        cols = df_processed.columns.tolist()
                        cols.remove(col)
                        idx = cols.index(after) + 1
                        cols.insert(idx, col)
                        df_processed = df_processed[cols]
                        execution_log.append(f"Moved column '{col}' after '{after}'")

                elif op_type == "create_summary":
                    group_cols = params.get("group_by", params.get("columns", []))
                    if group_cols:
                        valid = [c for c in group_cols if c in df_processed.columns]
                        if valid:
                            numeric = df_processed.select_dtypes(include=[np.number]).columns.tolist()
                            agg_cols = [c for c in numeric if c not in valid][:5]
                            if agg_cols:
                                summary_df = df_processed.groupby(valid)[agg_cols].mean().reset_index()
                                df_processed = summary_df
                                execution_log.append(f"Created summary grouped by {valid}")

                elif op_type == "clean_missing":
                    how = params.get("how", "any")
                    rows_before = len(df_processed)
                    candidate = df_processed.dropna(how=how)
                    # Guardrail: avoid wiping the whole dataset.
                    if rows_before > 0 and len(candidate) == 0:
                        execution_log.append(
                            f"Skipped clean_missing (how={how}): would remove all {rows_before} rows"
                        )
                    else:
                        df_processed = candidate
                        execution_log.append(f"Cleaned missing values: removed {rows_before - len(df_processed)} rows")

                elif op_type == "translate":
                    execution_log.append("Translate: requires LLM batch translation - not yet implemented")

                elif op_type == "filter_rows":
                    column = params.get("column")
                    operator = params.get("operator", "==")
                    value = params.get("value")
                    if column in df_processed.columns:
                        series = df_processed[column]
                        if operator == "==":
                            mask = series == value
                        elif operator == "!=":
                            mask = series != value
                        elif operator == ">":
                            mask = pd.to_numeric(series, errors="coerce") > pd.to_numeric(value, errors="coerce")
                        elif operator == ">=":
                            mask = pd.to_numeric(series, errors="coerce") >= pd.to_numeric(value, errors="coerce")
                        elif operator == "<":
                            mask = pd.to_numeric(series, errors="coerce") < pd.to_numeric(value, errors="coerce")
                        elif operator == "<=":
                            mask = pd.to_numeric(series, errors="coerce") <= pd.to_numeric(value, errors="coerce")
                        elif operator == "contains":
                            mask = series.astype(str).str.contains(str(value), case=False, na=False)
                        elif operator == "startswith":
                            mask = series.astype(str).str.startswith(str(value), na=False)
                        elif operator == "endswith":
                            mask = series.astype(str).str.endswith(str(value), na=False)
                        else:
                            mask = series == value
                        rows_before = len(df_processed)
                        df_processed = df_processed[mask.fillna(False)]
                        execution_log.append(
                            f"Filtered rows where {column} {operator} {value}: kept {len(df_processed)} of {rows_before}"
                        )

                elif op_type == "pivot_table":
                    group_by = params.get("group_by", [])
                    values = params.get("values", [])
                    aggfunc = params.get("aggfunc", "sum")
                    valid_group = [c for c in group_by if c in df_processed.columns]
                    valid_values = [c for c in values if c in df_processed.columns] if values else []
                    if valid_group:
                        if not valid_values:
                            valid_values = [
                                c for c in df_processed.select_dtypes(include=[np.number]).columns.tolist()
                                if c not in valid_group
                            ][:5]
                        if valid_values:
                            pivot_df = (
                                df_processed.pivot_table(
                                    index=valid_group,
                                    values=valid_values,
                                    aggfunc=aggfunc,
                                    fill_value=0,
                                )
                                .reset_index()
                            )
                            df_processed = pivot_df
                            execution_log.append(
                                f"Created pivot table by {valid_group} with {aggfunc} aggregation."
                            )

                elif op_type == "unpivot":
                    id_vars = params.get("id_vars", [])
                    value_vars = params.get("value_vars", [])
                    valid_id = [c for c in id_vars if c in df_processed.columns]
                    valid_value = [c for c in value_vars if c in df_processed.columns] if value_vars else None
                    var_name = params.get("var_name", "variable")
                    value_name = params.get("value_name", "value")
                    df_processed = df_processed.melt(
                        id_vars=valid_id,
                        value_vars=valid_value,
                        var_name=var_name,
                        value_name=value_name,
                    )
                    execution_log.append("Converted wide table to long format (unpivot).")

                elif op_type == "split_column":
                    column = params.get("column")
                    delimiter = params.get("delimiter", " ")
                    into = params.get("into", [])
                    if column in df_processed.columns:
                        split_df = df_processed[column].astype(str).str.split(delimiter, expand=True)
                        if into and isinstance(into, list):
                            for i, col_name in enumerate(into):
                                if i < split_df.shape[1]:
                                    df_processed[col_name] = split_df[i]
                        else:
                            for i in range(split_df.shape[1]):
                                df_processed[f"{column}_{i+1}"] = split_df[i]
                        execution_log.append(f"Split column '{column}' by delimiter '{delimiter}'.")

                elif op_type == "merge_columns":
                    cols = [c for c in params.get("columns", []) if c in df_processed.columns]
                    new_column = params.get("new_column")
                    separator = params.get("separator", " ")
                    if cols and new_column:
                        df_processed[new_column] = df_processed[cols].astype(str).agg(separator.join, axis=1)
                        execution_log.append(f"Merged columns {cols} into '{new_column}'.")

                elif op_type == "trim_text":
                    cols = [c for c in params.get("columns", []) if c in df_processed.columns]
                    for col in cols:
                        df_processed[col] = df_processed[col].astype(str).str.strip()
                    execution_log.append(f"Trimmed whitespace in columns: {cols}")

                elif op_type == "replace_values":
                    column = params.get("column")
                    mapping = params.get("mapping", {})
                    if column in df_processed.columns and isinstance(mapping, dict):
                        df_processed[column] = df_processed[column].replace(mapping)
                        execution_log.append(f"Replaced values in '{column}' using mapping.")

                elif op_type == "format_datetime":
                    cols = [c for c in params.get("columns", []) if c in df_processed.columns]
                    fmt = params.get("format", "%Y-%m-%d")
                    for col in cols:
                        dt_series = pd.to_datetime(df_processed[col], errors="coerce")
                        df_processed[col] = dt_series.dt.strftime(fmt)
                    execution_log.append(f"Formatted datetime columns {cols} as '{fmt}'.")

                elif op_type == "add_rank":
                    sort_by = params.get("sort_by")
                    new_column = params.get("new_column", "rank")
                    ascending = params.get("ascending", False)
                    if sort_by in df_processed.columns:
                        numeric = pd.to_numeric(df_processed[sort_by], errors="coerce")
                        df_processed[new_column] = numeric.rank(method="dense", ascending=ascending)
                        execution_log.append(f"Added rank column '{new_column}' by '{sort_by}'.")

                elif op_type == "running_total":
                    source = params.get("source_column")
                    new_col = params.get("new_column")
                    group_by = [c for c in params.get("group_by", []) if c in df_processed.columns]
                    if source in df_processed.columns and new_col:
                        numeric = pd.to_numeric(df_processed[source], errors="coerce").fillna(0)
                        if group_by:
                            df_processed[new_col] = numeric.groupby(
                                [df_processed[c] for c in group_by]
                            ).cumsum()
                        else:
                            df_processed[new_col] = numeric.cumsum()
                        execution_log.append(f"Added running total '{new_col}' from '{source}'.")

                elif op_type == "percentage_of_total":
                    source = params.get("source_column")
                    new_col = params.get("new_column")
                    group_by = [c for c in params.get("group_by", []) if c in df_processed.columns]
                    if source in df_processed.columns and new_col:
                        numeric = pd.to_numeric(df_processed[source], errors="coerce").fillna(0)
                        if group_by:
                            totals = numeric.groupby([df_processed[c] for c in group_by]).transform("sum")
                            totals = totals.replace(0, np.nan)
                            df_processed[new_col] = (numeric / totals) * 100
                        else:
                            total = numeric.sum()
                            df_processed[new_col] = (numeric / total * 100) if total else 0
                        execution_log.append(f"Added percentage-of-total column '{new_col}'.")

                elif op_type == "clip_outliers":
                    cols = [c for c in params.get("columns", []) if c in df_processed.columns]
                    lower_q = params.get("lower_quantile", 0.01)
                    upper_q = params.get("upper_quantile", 0.99)
                    for col in cols:
                        numeric = pd.to_numeric(df_processed[col], errors="coerce")
                        if numeric.notna().any():
                            lower = numeric.quantile(lower_q)
                            upper = numeric.quantile(upper_q)
                            df_processed[col] = numeric.clip(lower=lower, upper=upper)
                    execution_log.append(
                        f"Clipped outliers in {cols} using quantiles [{lower_q}, {upper_q}]."
                    )

            except Exception as e:
                print(f"Error executing operation {op_type}: {e}")
                execution_log.append(f"Error in {op_type}: {str(e)}")
        
        return df_processed, execution_log

    @staticmethod
    def _to_serializable_value(value: Any) -> Any:
        """Convert numpy/pandas values to JSON-safe primitives."""
        if pd.isna(value):
            return None
        if isinstance(value, (np.integer,)):
            return int(value)
        if isinstance(value, (np.floating,)):
            if np.isnan(value) or np.isinf(value):
                return None
            return float(value)
        if isinstance(value, (pd.Timestamp,)):
            return value.isoformat()
        return value

    @staticmethod
    def build_visual_pack(
        df: pd.DataFrame,
        user_query: str = "",
        dashboard_template: str = "operations",
    ) -> Dict[str, Any]:
        """
        Build lightweight dashboard/diagram payload from current dataframe.
        Safe, deterministic, and small enough for chat response rendering.
        """
        q = (user_query or "").lower()
        template = (dashboard_template or "operations").lower().strip()
        if template not in {"executive", "sales", "operations", "finance"}:
            template = "operations"
        asked_for_visual = any(
            token in q for token in ["diagram", "dashboard", "visual", "chart", "graph", "plot"]
        )

        total_rows = int(df.shape[0])
        total_columns = int(df.shape[1])
        missing_values = int(df.isna().sum().sum())
        duplicate_rows = int(df.duplicated().sum())

        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        categorical_cols = df.select_dtypes(include=["object", "category"]).columns.tolist()

        kpis = [
            {"label": "Rows", "value": total_rows},
            {"label": "Columns", "value": total_columns},
            {"label": "Missing Values", "value": missing_values},
            {"label": "Duplicate Rows", "value": duplicate_rows},
        ]
        if template == "executive":
            kpis = [
                {"label": "Total Records", "value": total_rows},
                {"label": "Data Fields", "value": total_columns},
                {"label": "Completeness Gaps", "value": missing_values},
                {"label": "Duplicate Risk", "value": duplicate_rows},
            ]
        elif template == "sales":
            kpis = [
                {"label": "Sales Rows", "value": total_rows},
                {"label": "Dimensions", "value": total_columns},
                {"label": "Missing Cells", "value": missing_values},
                {"label": "Repeated Orders", "value": duplicate_rows},
            ]
        elif template == "finance":
            kpis = [
                {"label": "Ledger Rows", "value": total_rows},
                {"label": "Attributes", "value": total_columns},
                {"label": "Null Entries", "value": missing_values},
                {"label": "Duplicate Entries", "value": duplicate_rows},
            ]

        charts: List[Dict[str, Any]] = []

        if categorical_cols:
            cat_col = categorical_cols[0]
            if template in {"sales", "executive"}:
                for candidate in categorical_cols:
                    col = str(candidate).lower()
                    if any(t in col for t in ["region", "area", "city", "category", "segment", "shop"]):
                        cat_col = candidate
                        break
            top_counts = (
                df[cat_col]
                .fillna("Missing")
                .astype(str)
                .value_counts()
                .head(8)
            )
            charts.append(
                {
                    "title": f"{template.title()} view: top values in {cat_col}",
                    "type": "bar",
                    "x_key": "category",
                    "y_key": "count",
                    "data": [
                        {"category": str(idx), "count": int(val)}
                        for idx, val in top_counts.items()
                    ],
                }
            )

        if numeric_cols:
            num_col = numeric_cols[0]
            if template in {"sales", "finance", "executive"}:
                for candidate in numeric_cols:
                    col = str(candidate).lower()
                    if any(t in col for t in ["sales", "revenue", "amount", "profit", "cost", "total"]):
                        num_col = candidate
                        break
            trend_sample = df[[num_col]].copy().head(30).reset_index(drop=True)
            trend_sample["row"] = trend_sample.index + 1
            charts.append(
                {
                    "title": f"{template.title()} trend: {num_col} (first 30 rows)",
                    "type": "line",
                    "x_key": "row",
                    "y_key": "value",
                    "data": [
                        {
                            "row": int(row["row"]),
                            "value": DataService._to_serializable_value(row[num_col]),
                        }
                        for _, row in trend_sample.iterrows()
                    ],
                }
            )

        summary_lines = []
        if asked_for_visual:
            summary_lines.append(f"{template.title()} dashboard and diagram assets prepared from your latest dataset.")
        summary_lines.append(
            f"Generated {len(charts)} chart-ready panel(s) with {len(kpis)} KPI cards."
        )

        return {
            "requested": asked_for_visual,
            "template": template,
            "summary": " ".join(summary_lines),
            "kpis": kpis,
            "charts": charts,
        }

    @staticmethod
    def build_dashboard_excel_outputs(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        Build additional dashboard-style Excel sheets for generated outputs.
        """
        outputs: Dict[str, pd.DataFrame] = {}
        kpi_df = pd.DataFrame(
            [
                {"Metric": "Rows", "Value": int(df.shape[0])},
                {"Metric": "Columns", "Value": int(df.shape[1])},
                {"Metric": "Missing Values", "Value": int(df.isna().sum().sum())},
                {"Metric": "Duplicate Rows", "Value": int(df.duplicated().sum())},
            ]
        )
        outputs["KPI_Summary"] = kpi_df

        categorical_cols = df.select_dtypes(include=["object", "category"]).columns.tolist()
        if categorical_cols:
            cat_col = categorical_cols[0]
            top_df = (
                df[cat_col]
                .fillna("Missing")
                .astype(str)
                .value_counts()
                .head(15)
                .reset_index()
            )
            top_df.columns = [cat_col, "Count"]
            outputs["Top_Categories"] = top_df

        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        if numeric_cols:
            stats_df = df[numeric_cols].describe().transpose().reset_index()
            stats_df.rename(columns={"index": "Column"}, inplace=True)
            outputs["Numeric_Stats"] = stats_df

        return outputs

    @staticmethod
    def table_page(
        df: pd.DataFrame,
        offset: int = 0,
        limit: int = 200,
    ) -> Dict[str, Any]:
        """
        Return paginated, JSON-safe table rows for UI spreadsheet rendering.
        """
        safe_offset = max(0, int(offset or 0))
        safe_limit = max(1, min(int(limit or 200), 1000))
        page_df = df.iloc[safe_offset:safe_offset + safe_limit].copy()
        records = DataService._sanitize_for_json(page_df.to_dict(orient="records"))
        return {
            "offset": safe_offset,
            "limit": safe_limit,
            "total_rows": int(df.shape[0]),
            "columns": [str(c) for c in df.columns.tolist()],
            "rows": records,
        }

    @staticmethod
    def build_auto_plan(
        profile: Dict[str, Any],
        target_column: str,
        task: str,
        feature_types: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """
        Build a data wrangling plan from FTI and prediction engineering (AutoDW-style).
        - Obligatory cleaning: drop constant columns, drop NA in target, remove duplicates.
        - Enrichment by feature type: normalize numeric, encode categorical, datetime, vectorize text.
        """
        plan = []
        columns = profile.get("columns", [])
        numeric_cols = profile.get("numeric_columns", [])
        text_cols = profile.get("text_columns", [])
        categorical_cols = profile.get("categorical_columns", [])

        # Obligatory cleaning
        plan.append({"type": "drop_constant_columns", "params": {}})
        if target_column and target_column in columns:
            plan.append({"type": "drop_na", "params": {"columns": [target_column]}})
        plan.append({"type": "remove_duplicates", "params": {}})

        # Drop ignorable ID columns
        ignorable = [c for c, ft in feature_types.items() if ft == "ignorable_id"]
        if ignorable:
            plan.append({"type": "drop_columns", "params": {"columns": ignorable}})

        # Enrichment: collect columns by type then add one op per type
        num_to_norm = [c for c, ft in feature_types.items() if ft == "numerical" and c in numeric_cols and c in columns]
        cat_to_enc = [c for c, ft in feature_types.items() if ft == "categorical" and c in categorical_cols and c in columns]
        dt_cols = [c for c, ft in feature_types.items() if ft == "datetime" and c in columns]
        sentence_cols = [c for c, ft in feature_types.items() if ft == "sentence" and c in text_cols and c in columns]
        embedded_cols = [c for c, ft in feature_types.items() if ft == "embedded_number" and c in columns]

        for col in embedded_cols:
            plan.append({"type": "convert_type", "params": {"columns": [col], "target_type": "float"}})
        if num_to_norm or embedded_cols:
            plan.append({"type": "normalize_numeric", "params": {"columns": num_to_norm + embedded_cols}})
        if cat_to_enc:
            plan.append({"type": "encode_categorical", "params": {"columns": cat_to_enc}})
        if dt_cols:
            plan.append({"type": "extract_datetime_features", "params": {"columns": dt_cols}})
        for col in sentence_cols:
            plan.append({"type": "vectorize_text", "params": {"columns": [col], "max_features": 50}})

        plan.append({"type": "fill_na", "params": {"method": "mean"}})
        return plan
