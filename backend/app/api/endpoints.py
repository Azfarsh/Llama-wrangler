from fastapi import APIRouter, UploadFile, File, HTTPException, Body
from fastapi.responses import FileResponse, PlainTextResponse
from app.services.data_service import DataService
from app.services.llm_service import LLMService
from app.services.rag_service import RAGService
from app.services.fti_service import infer_dataset_fti
from app.services.code_generator import generate_wrangling_script
from app.services.excel_ai_service import ExcelAIService
from app.schemas import validate_plan
from app.core.config import settings
import os
import pandas as pd
import uuid
import json
from openpyxl import load_workbook

router = APIRouter()
data_service = DataService()
llm_service = LLMService()
rag_service = RAGService()
excel_ai_service = ExcelAIService()

# In-memory storage for session state (prototype only)
# In production, use a database or Redis
sessions = {}
excel_sessions = {}

VISUAL_QUERY_TOKENS = ("diagram", "dashboard", "visual", "chart", "graph", "plot")
TRANSFORM_QUERY_TOKENS = (
    "clean", "remove", "drop", "fill", "replace", "rename", "convert", "filter",
    "split", "merge", "sort", "pivot", "unpivot", "rank", "running total",
    "percentage", "clip", "format", "deduplicate", "trim", "add column",
)


def _get_df(session_id: str):
    """Load and return single DataFrame for session."""
    current_file_path = sessions[session_id].get("current_file_path") or sessions[session_id]["file_path"]
    sheet_name = sessions[session_id].get("current_sheet_name", sessions[session_id].get("sheet_name"))
    data = data_service.load_dataset(current_file_path, sheet_name)
    return data_service.get_single_df(data, sheet_name)


def _wants_visual_outputs(user_query: str) -> bool:
    q = (user_query or "").lower()
    return any(token in q for token in VISUAL_QUERY_TOKENS)


def _wants_transformations(user_query: str) -> bool:
    q = (user_query or "").lower()
    return any(token in q for token in TRANSFORM_QUERY_TOKENS)


@router.post("/upload")
async def upload_dataset(file: UploadFile = File(...)):
    session_id = str(uuid.uuid4())
    file_ext = os.path.splitext(file.filename)[1].lower()
    file_path = os.path.join(settings.UPLOAD_DIR, f"{session_id}{file_ext}")
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    try:
        data = data_service.load_dataset(file_path)
        df = data_service.get_single_df(data)
        profile = data_service.profile_dataset(df)
        profile["session_id"] = session_id
        profile["sheet_names"] = list(data.keys()) if isinstance(data, dict) else ["Sheet1"]
        
        sessions[session_id] = {
            "file_path": file_path,
            "filename": file.filename,
            "file_ext": file_ext,
            "sheet_name": None,
            "current_file_path": file_path,
            "current_sheet_name": None,
            "sheet_names": profile.get("sheet_names", ["Sheet1"]),
            "conversation_history": [],
        }
        rag_service.index_profile(profile, session_id)
        return {
            "session_id": session_id,
            "filename": file.filename,
            "profile": profile,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"Upload failed: {str(e)}")

@router.post("/analyze")
async def analyze_intent(request: dict = Body(...)):
    session_id = request.get("session_id")
    user_query = request.get("query")
    
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    df = _get_df(session_id)
    profile = data_service.profile_dataset(df)
    profile["session_id"] = session_id
    sessions[session_id]["profile"] = profile
    conversation_history = sessions[session_id].get("conversation_history", [])
    if user_query:
        conversation_history.append({"role": "user", "text": user_query})
        sessions[session_id]["conversation_history"] = conversation_history[-20:]
    
    # RAG: get relevant context for query
    rag_context = rag_service.get_rag_context(user_query, profile)
    
    # Agent 1: Intent Understanding (with RAG context)
    intent = await llm_service.analyze_intent(user_query, profile, rag_context)
    
    # Agent 2: Planning Agent (with RAG context)
    plan = await llm_service.generate_plan(user_query, profile, intent, rag_context)
    
    # Validate plan - schema check, no arbitrary code
    plan, plan_warnings = validate_plan(plan, available_columns=list(df.columns))
    
    return {
        "intent": intent,
        "plan": plan,
        "plan_warnings": plan_warnings,
        "profile": profile
    }

@router.post("/execute")
async def execute_plan(request: dict = Body(...)):
    session_id = request.get("session_id")
    plan = request.get("plan")
    intent = request.get("intent", {})
    user_query = (request.get("query") or "").strip()
    force_visual_outputs = bool(request.get("force_visual_outputs", False))
    fast_mode = bool(request.get("fast_mode", False))
    dashboard_template = str(request.get("dashboard_template", "")).strip().lower()
    if dashboard_template not in {"executive", "sales", "operations", "finance"}:
        dashboard_template = "operations"
    
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    df = _get_df(session_id)
    original_profile = data_service.profile_dataset(df)
    conversation_history = sessions[session_id].get("conversation_history", [])
    
    # Validate plan before execution (used as fallback / transparent plan output)
    plan, plan_warnings = validate_plan(plan, available_columns=list(df.columns))

    processed_df = None
    execution_log = []
    generated_code = None
    execution_mode = "plan"
    extra_excel_sheets = {}
    generation_errors = []
    wants_visual_outputs = force_visual_outputs or _wants_visual_outputs(user_query)
    wants_transformations = _wants_transformations(user_query)
    visual_only_request = wants_visual_outputs and not wants_transformations and not plan

    if visual_only_request:
        processed_df = df.copy()
        execution_mode = "visual_fast_path"
        execution_log = ["Fast path: visual dashboard generation without LLM transformation."]
    elif user_query:
        rag_context = rag_service.get_rag_context(user_query, original_profile)
        max_attempts = 3
        for attempt in range(max_attempts):
            previous_error = generation_errors[-1] if generation_errors else None
            try:
                generated_code = await llm_service.generate_dataframe_code(
                    user_query=user_query,
                    dataset_metadata=original_profile,
                    rag_context=rag_context,
                    previous_error=previous_error,
                    conversation_history=conversation_history,
                )
                processed_df, execution_log, extra_excel_sheets = data_service.execute_generated_code(df, generated_code)
                # If query is transformational but result is unchanged, force a retry with stronger feedback.
                requires_transformation = bool(intent.get("requires_transformation", True))
                if requires_transformation and processed_df.equals(df):
                    raise ValueError("Generated code executed but dataset remained unchanged. Apply the requested transformation.")
                execution_mode = "generated_code"
                execution_log = execution_log + [f"Execution mode: generated_code (attempt {attempt + 1})"]
                break
            except Exception as e:
                generation_errors.append(str(e))

    if processed_df is None:
        processed_df, execution_log = data_service.execute_plan(df, plan)
        if generation_errors:
            execution_log = execution_log + [f"Code generation fallback reason: {generation_errors[-1]}"]
        execution_log = execution_log + ["Execution mode: validated_plan"]

    if wants_visual_outputs:
        dashboard_sheets = data_service.build_dashboard_excel_outputs(processed_df)
        # Preserve sheets produced by generated code and add deterministic dashboard sheets.
        extra_excel_sheets = {**extra_excel_sheets, **dashboard_sheets}
        execution_log.append("Generated dashboard-ready sheets for visual analysis.")

    processed_profile = data_service.profile_dataset(processed_df)
    if fast_mode:
        insights = llm_service._build_deterministic_insights(original_profile, processed_profile, execution_log)
    else:
        insights = await llm_service.generate_insights(original_profile, processed_profile, execution_log, intent)
    visual_pack = data_service.build_visual_pack(processed_df, user_query, dashboard_template)
    
    file_ext = sessions[session_id].get("file_ext", ".csv")
    output_filename = f"{session_id}_processed{file_ext}"
    output_path = os.path.join(settings.UPLOAD_DIR, output_filename)
    
    if file_ext in (".xlsx", ".xls"):
        data_service.save_excel_with_outputs(output_path, processed_df, extra_excel_sheets)
        sessions[session_id]["current_sheet_name"] = "TransformedData"
    else:
        processed_df.to_csv(output_path, index=False)
        sessions[session_id]["current_sheet_name"] = None

    # Make the processed output the new working dataset for follow-up prompts.
    sessions[session_id]["current_file_path"] = output_path
    
    sessions[session_id]["processed_df_path"] = output_path
    sessions[session_id]["insights"] = insights
    sessions[session_id]["execution_log"] = execution_log
    sessions[session_id]["plan"] = plan
    sessions[session_id]["generated_code"] = generated_code
    sessions[session_id]["execution_mode"] = execution_mode
    sessions[session_id]["extra_excel_sheets"] = list(extra_excel_sheets.keys())
    sessions[session_id]["visual_pack"] = visual_pack
    sessions[session_id]["profile"] = processed_profile
    sessions[session_id]["conversation_history"] = (
        conversation_history
        + [
            {"role": "assistant", "text": f"Execution mode: {execution_mode}"},
            {"role": "assistant", "text": " | ".join(execution_log[:5]) if execution_log else "No execution steps"},
        ]
    )[-20:]

    preview = processed_profile.get("head", [])
    
    return {
        "message": "Execution successful",
        "preview": preview,
        "profile": processed_profile,
        "insights": insights,
        "execution_log": execution_log,
        "plan_warnings": plan_warnings,
        "execution_mode": execution_mode,
        "generation_warnings": list(dict.fromkeys(generation_errors)),
        "extra_excel_sheets": list(extra_excel_sheets.keys()),
        "visual_pack": visual_pack,
        "download_url": f"/api/download/{output_filename}"
    }

@router.get("/session/{session_id}/insights")
async def get_insights(session_id: str):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    insights = sessions[session_id].get("insights", {})
    return insights


@router.get("/session/{session_id}/table")
async def get_table_data(session_id: str, offset: int = 0, limit: int = 200):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    df = _get_df(session_id)
    return data_service.table_page(df, offset=offset, limit=limit)


@router.post("/wrangle")
async def auto_wrangle(request: dict = Body(...)):
    """
    Full AutoDW-style pipeline: predict target + FTI -> build plan -> execute -> return results + code.
    Optional: user_query to override or guide; if omitted, runs fully automatic.
    """
    session_id = request.get("session_id")
    user_query = request.get("query", "").strip()
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    df = _get_df(session_id)
    profile = data_service.profile_dataset(df)
    sessions[session_id]["profile"] = profile

    # 1) Prediction engineering: recommend target column and task
    columns = profile.get("columns", [])
    target_result = await llm_service.recommend_target_column(columns)
    target_column = target_result.get("target_column") or (columns[-1] if columns else "")
    task = target_result.get("task", "classification")
    n_unique = int(df[target_column].nunique()) if target_column in df.columns else 0
    if n_unique > 30 and df[target_column].dtype in [float, "float64"]:
        task = "regression"
    else:
        task = "classification"

    # 2) Feature type inference (rule-based)
    feature_types = infer_dataset_fti(df)

    # 3) Build plan from FTI + target
    plan = data_service.build_auto_plan(profile, target_column, task, feature_types)
    if user_query:
        rag_context = rag_service.get_rag_context(user_query, profile)
        custom_plan = await llm_service.generate_plan(user_query, profile, {"goal": "custom", "key_requirements": [user_query]}, rag_context)
        if custom_plan and len(custom_plan) > 0:
            plan = custom_plan

    plan, plan_warnings = validate_plan(plan)
    processed_df, execution_log = data_service.execute_plan(df, plan)
    processed_profile = data_service.profile_dataset(processed_df)
    insights = await llm_service.generate_insights(profile, processed_profile, execution_log, {"goal": "Auto wrangle"})

    file_ext = sessions[session_id].get("file_ext", ".csv")
    output_filename = f"{session_id}_processed{file_ext}"
    output_path = os.path.join(settings.UPLOAD_DIR, output_filename)
    if file_ext in (".xlsx", ".xls"):
        processed_df.to_excel(output_path, index=False)
        sessions[session_id]["current_sheet_name"] = None
    else:
        processed_df.to_csv(output_path, index=False)
        sessions[session_id]["current_sheet_name"] = None

    sessions[session_id]["processed_df_path"] = output_path
    sessions[session_id]["current_file_path"] = output_path
    sessions[session_id]["insights"] = insights
    sessions[session_id]["execution_log"] = execution_log
    sessions[session_id]["plan"] = plan
    sessions[session_id]["target_column"] = target_column
    sessions[session_id]["task"] = task
    sessions[session_id]["feature_types"] = feature_types

    return {
        "message": "Auto wrangling complete",
        "target_column": target_column,
        "task": task,
        "feature_types": feature_types,
        "plan": plan,
        "plan_warnings": plan_warnings,
        "preview": processed_profile.get("head", []),
        "profile": processed_profile,
        "insights": insights,
        "execution_log": execution_log,
        "download_url": f"/api/download/{output_filename}",
    }


@router.get("/code/{session_id}")
async def get_generated_code(session_id: str):
    """Return generated Python script for the wrangling session."""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    s = sessions[session_id]
    file_path = s.get("file_path")
    plan = s.get("plan")
    if not file_path or not plan:
        raise HTTPException(status_code=400, detail="No plan available; run execute or wrangle first.")
    output_path = s.get("processed_df_path") or os.path.join(settings.UPLOAD_DIR, f"{session_id}_processed.csv")
    script = generate_wrangling_script(file_path, output_path, plan)
    return PlainTextResponse(script, media_type="text/plain")


@router.get("/download/{filename}")
async def download_file(filename: str):
    file_path = os.path.join(settings.UPLOAD_DIR, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, filename=filename)
    raise HTTPException(status_code=404, detail="File not found")


@router.post("/excel/upload")
async def upload_excel_file(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")
    session_id = str(uuid.uuid4())
    file_path = os.path.join(settings.UPLOAD_DIR, f"{session_id}.xlsx")
    with open(file_path, "wb") as buffer:
        buffer.write(await file.read())
    try:
        workbook = load_workbook(file_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")

    sheet_json = excel_ai_service.workbook_to_sheet_json(workbook)
    sheet_json, _ = excel_ai_service.enrich_sheet_json_and_changes_from_file(
        file_path, sheet_json, None
    )
    excel_sessions[session_id] = {
        "filename": file.filename,
        "original_path": file_path,
        "current_path": file_path,
    }
    return {
        "session_id": session_id,
        "filename": file.filename,
        "sheet_names": workbook.sheetnames,
        "sheet_json": sheet_json,
    }


@router.post("/excel/ai")
async def process_excel_ai(request: dict = Body(...)):
    session_id = request.get("session_id")
    user_message = (request.get("message") or "").strip()
    if not session_id or session_id not in excel_sessions:
        raise HTTPException(status_code=404, detail="Please upload an Excel file first.")
    if not user_message:
        raise HTTPException(status_code=400, detail="Please provide a message for AI processing.")

    if not settings.GEMINI_API_KEY:
        raise HTTPException(status_code=400, detail="Gemini is not configured. Set GEMINI_API_KEY in backend environment.")

    current_path = excel_sessions[session_id]["current_path"]
    try:
        workbook = load_workbook(current_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load workbook: {e}")

    try:
        result = excel_ai_service.process_excel_request(
            workbook=workbook,
            user_message=user_message,
            input_sheet_json=request.get("sheet_json"),
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI processing failed: {e}")

    is_read_only = result.get("readOnly", False)

    if is_read_only:
        return {
            "message": "AI analyzed your spreadsheet.",
            "result": result,
            "updated_sheet_json": None,
            "changes_made": [],
            "charts_created": [],
            "dashboard_summary": {"created": False, "sheet": None, "elements": 0},
            "changed_sheets": [],
            "download_url": None,
        }

    updated_path = os.path.join(settings.UPLOAD_DIR, f"{session_id}_updated.xlsx")
    workbook.save(updated_path)
    excel_sessions[session_id]["current_path"] = updated_path

    updated_sheet_json = excel_ai_service.workbook_to_sheet_json(workbook)
    change_tracking = list(result.get("change_tracking", []))
    updated_sheet_json, change_tracking = excel_ai_service.enrich_sheet_json_and_changes_from_file(
        updated_path, updated_sheet_json, change_tracking
    )
    result["change_tracking"] = change_tracking
    dashboard_summary = result.get("dashboard_summary", {})

    return {
        "message": "AI modified your spreadsheet.",
        "result": result,
        "updated_sheet_json": updated_sheet_json,
        "changes_made": change_tracking,
        "charts_created": result.get("applied_charts", []),
        "dashboard_summary": dashboard_summary,
        "changed_sheets": result.get("changed_sheets", []),
        "download_url": f"/api/excel/download?session_id={session_id}",
    }


@router.get("/excel/download")
async def download_excel_file(session_id: str):
    if session_id not in excel_sessions:
        raise HTTPException(status_code=404, detail="Session not found.")
    path = excel_sessions[session_id].get("current_path")
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Updated file not found.")
    filename = os.path.basename(path)
    return FileResponse(path, filename=filename)


@router.get("/excel/health")
async def excel_ai_health(check_model: bool = False):
    key_hint = (settings.GEMINI_API_KEY[:8] + "...") if settings.GEMINI_API_KEY else ""
    has_key = bool(settings.GEMINI_API_KEY and settings.GEMINI_API_KEY != "PASTE_YOUR_GEMINI_KEY_HERE")
    if not has_key:
        return {"ok": False, "configured": False, "detail": "Set GEMINI_API_KEY in backend/.env", "key_hint": key_hint}
    if not check_model:
        return {"ok": True, "configured": True, "model": settings.GEMINI_MODEL, "key_hint": key_hint}
    try:
        response = excel_ai_service.client.models.generate_content(
            model=settings.GEMINI_MODEL,
            contents="Reply with: Gemini health check passed",
        )
        return {"ok": True, "configured": True, "model": settings.GEMINI_MODEL, "response": (response.text or "").strip(), "key_hint": key_hint}
    except Exception as e:
        return {"ok": False, "configured": True, "model": settings.GEMINI_MODEL, "detail": str(e), "key_hint": key_hint}
